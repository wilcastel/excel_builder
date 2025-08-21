# -*- coding: utf-8 -*-
"""
Gestor de exportación de archivos Excel - Versión Optimizada
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Callable
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config.settings import AppSettings
from models.column_config import ColumnConfig, DataType

class ExportManager:
    """Gestor para exportación de archivos Excel con formato - Versión Optimizada"""
    
    def __init__(self, settings: AppSettings):
        self.settings = settings
        self.workbook = None
        self.worksheet = None
        self.mapping_config = None
        self.mapping_manager = None
        # Inicializar el generador numérico inmediatamente
        from core.numeric_generator import NumericGenerator
        self.numeric_generator = NumericGenerator()
        self.logger = logging.getLogger(__name__)
        
        # Cache para optimización
        self._mapping_cache = {}
        self._date_format_cache = {}
        self._column_letter_cache = {}
        
    def _preprocess_numeric_groups(self, data: pd.DataFrame, col_config: ColumnConfig):
        """Pre-procesar grupos para asignar números correctamente"""
        if not hasattr(col_config, 'numeric_grouping_columns') or not col_config.numeric_grouping_columns:
            return
        
        # Limpiar contadores existentes
        self.numeric_generator.counters.clear()
        
        # Procesar todos los datos para identificar grupos únicos
        unique_groups = set()
        for _, row in data.iterrows():
            group_values = []
            for group_col in col_config.numeric_grouping_columns:
                # Buscar la columna con diferentes variaciones del nombre
                found_col = None
                
                # 1. Búsqueda exacta
                if group_col in row.index:
                    found_col = group_col
                # 2. Búsqueda por nombre en minúsculas
                elif group_col.lower() in [col.lower() for col in row.index]:
                    for col in row.index:
                        if col.lower() == group_col.lower():
                            found_col = col
                            break
                # 3. Búsqueda por palabras clave específicas
                else:
                    # Mapeo de nombres comunes
                    name_mappings = {
                        'nombre programa': ['nombre', 'programa'],
                        'fecha': ['fecha'],
                        'fecha (dd/mm/aa)': ['fecha'],
                        'tema': ['tema', 'nombre', 'programa'],
                        'programa': ['programa', 'nombre'],
                        'nombre': ['nombre', 'programa']
                    }
                    
                    # Buscar en el mapeo
                    for key, keywords in name_mappings.items():
                        if any(keyword in group_col.lower() for keyword in [key] + keywords):
                            # Buscar columnas que contengan estas palabras clave
                            for col in row.index:
                                col_lower = col.lower()
                                if any(keyword in col_lower for keyword in keywords):
                                    found_col = col
                                    break
                            if found_col:
                                break
                
                # 4. Búsqueda por similitud (último recurso)
                if not found_col:
                    for col in row.index:
                        # Si más del 70% de las palabras coinciden
                        group_words = set(group_col.lower().split())
                        col_words = set(col.lower().split())
                        if group_words and col_words:
                            similarity = len(group_words.intersection(col_words)) / len(group_words.union(col_words))
                            if similarity > 0.3:  # 30% de similitud
                                found_col = col
                                break
                
                if found_col:
                    group_value = str(row[found_col])
                    group_values.append(group_value)
            
            group_key = '|'.join(group_values) if group_values else 'default'
            unique_groups.add(group_key)
        
        # Asignar números a cada grupo único
        start_value = getattr(col_config, 'numeric_start', 1) or 1
        for i, group_key in enumerate(sorted(unique_groups)):
            self.numeric_generator.counters[group_key] = start_value + i
        
        self.logger.info(f"Pre-procesados {len(unique_groups)} grupos únicos para generador numérico")
    
    def _get_column_value(self, row: pd.Series, col_config: ColumnConfig, full_data: pd.DataFrame = None) -> Any:
        """Obtener valor de columna según configuración, aplicando mapeos y generadores si están configurados"""
        
        # Si es un generador numérico, generar valor
        if hasattr(col_config, 'is_numeric_generator') and col_config.is_numeric_generator:
            try:
                # Si hay columnas de agrupación, usar generación agrupada
                if hasattr(col_config, 'numeric_grouping_columns') and col_config.numeric_grouping_columns and full_data is not None:
                    # Crear clave de grupo basada en los valores de las columnas de agrupación
                    group_values = []
                    for group_col in col_config.numeric_grouping_columns:
                        # Buscar la columna con diferentes variaciones del nombre
                        found_col = None
                        
                        # 1. Búsqueda exacta
                        if group_col in row.index:
                            found_col = group_col
                        # 2. Búsqueda por nombre en minúsculas
                        elif group_col.lower() in [col.lower() for col in row.index]:
                            for col in row.index:
                                if col.lower() == group_col.lower():
                                    found_col = col
                                    break
                        # 3. Búsqueda por palabras clave específicas
                        else:
                            # Mapeo de nombres comunes
                            name_mappings = {
                                'nombre programa': ['nombre', 'programa'],
                                'fecha': ['fecha'],
                                'fecha (dd/mm/aa)': ['fecha'],
                                'tema': ['tema', 'nombre', 'programa'],
                                'programa': ['programa', 'nombre'],
                                'nombre': ['nombre', 'programa']
                            }
                            
                            # Buscar en el mapeo
                            for key, keywords in name_mappings.items():
                                if any(keyword in group_col.lower() for keyword in [key] + keywords):
                                    # Buscar columnas que contengan estas palabras clave
                                    for col in row.index:
                                        col_lower = col.lower()
                                        if any(keyword in col_lower for keyword in keywords):
                                            found_col = col
                                            break
                                    if found_col:
                                        break
                        
                        # 4. Búsqueda por similitud (último recurso)
                        if not found_col:
                            for col in row.index:
                                # Si más del 70% de las palabras coinciden
                                group_words = set(group_col.lower().split())
                                col_words = set(col.lower().split())
                                if group_words and col_words:
                                    similarity = len(group_words.intersection(col_words)) / len(group_words.union(col_words))
                                    if similarity > 0.3:  # 30% de similitud
                                        found_col = col
                                        break
                        
                        if found_col:
                            group_value = str(row[found_col])
                            group_values.append(group_value)
                    
                    group_key = '|'.join(group_values) if group_values else 'default'
                    
                    # Usar el contador para este grupo específico
                    if group_key not in self.numeric_generator.counters:
                        start_value = getattr(col_config, 'numeric_start', 1) or 1
                        self.numeric_generator.counters[group_key] = start_value
                    
                    # NO incrementar el contador aquí - mantener el mismo número para el grupo
                    return int(self.numeric_generator.counters[group_key])
                else:
                    # Generar secuencial simple
                    if 'simple' not in self.numeric_generator.counters:
                        start_value = getattr(col_config, 'numeric_start', 1) or 1
                        self.numeric_generator.counters['simple'] = start_value
                    else:
                        self.numeric_generator.counters['simple'] += 1
                    
                    return int(self.numeric_generator.counters['simple'])
                    
            except Exception as e:
                self.logger.error(f"Error generando número para columna '{col_config.display_name}': {e}")
                return getattr(col_config, 'numeric_start', 1) or 1
        
        # Si tiene configuración de mapeo dinámico, aplicar mapeo
        if (hasattr(col_config, 'mapping_source') and col_config.mapping_source and 
            hasattr(col_config, 'mapping_key_column') and col_config.mapping_key_column and 
            hasattr(col_config, 'mapping_value_column') and col_config.mapping_value_column and 
            self.mapping_config):
            
            try:
                # Obtener valor de la columna fuente
                if col_config.mapping_source in row.index:
                    source_value = str(row[col_config.mapping_source]).strip()
                    
                    # Normalizar fechas si la columna fuente es 'Fecha'
                    if col_config.mapping_source.lower() in ['fecha', 'date']:
                        try:
                            # Convertir datetime a formato dd/mm/yy
                            from datetime import datetime
                            if ' ' in source_value:  # Si tiene hora
                                dt = datetime.strptime(source_value, '%Y-%m-%d %H:%M')
                            else:
                                dt = datetime.strptime(source_value, '%Y-%m-%d')
                            source_value = dt.strftime('%d/%m/%y')
                        except:
                            pass  # Si falla la conversión, usar el valor original
                    
                    # Crear identificador único para este mapeo
                    mapping_id = f"{col_config.mapping_source}_{col_config.mapping_key_column}_{col_config.mapping_value_column}"
                    
                    # Usar cache para mapeos
                    cache_key = f"{mapping_id}_{source_value}"
                    if cache_key in self._mapping_cache:
                        return self._mapping_cache[cache_key]
                    
                    # Buscar en la configuración de mapeo
                    if mapping_id in self.mapping_config:
                        mapping_info = self.mapping_config[mapping_id]
                        mapping_data = mapping_info['mapping_dict']
                        additional_keys = mapping_info.get('additional_keys', [])
                        
                        # Crear clave de búsqueda
                        if additional_keys:
                            # Clave compuesta con columnas adicionales
                            search_parts = [source_value]
                            
                            # Agregar valores de columnas adicionales del archivo fuente
                            for additional_col in additional_keys:
                                if additional_col in row.index:
                                    search_parts.append(str(row[additional_col]).strip())
                                else:
                                    search_parts.append("")
                            
                            search_key = "|".join(search_parts)
                        else:
                            # Búsqueda simple
                            search_key = source_value
                        
                        # Buscar el valor en el mapeo (comparación exacta primero, luego case-insensitive)
                        if search_key in mapping_data:
                            result = mapping_data[search_key]
                            self._mapping_cache[cache_key] = result
                            return result
                        
                        # Buscar con comparación case-insensitive
                        for key, value in mapping_data.items():
                            if str(key).strip().lower() == search_key.lower():
                                result = value
                                self._mapping_cache[cache_key] = result
                                return result
                        
                        # Si no se encuentra, devolver valor original
                        self._mapping_cache[cache_key] = source_value
                        return source_value
                    else:
                        # Si no hay configuración de mapeo, devolver vacío
                        return ''
                else:
                    return ''
                    
            except Exception as e:
                self.logger.error(f"Error aplicando mapeo para columna '{col_config.display_name}': {e}")
                return row.get(col_config.mapping_source, '')
        
        # Si no es generador ni mapeo, obtener valor de columna fuente
        if hasattr(col_config, 'source_column') and col_config.source_column and col_config.source_column in row.index:
            return row[col_config.source_column]
        
        # Si no hay columna fuente, devolver vacío
        return ''
    
    def _apply_dynamic_mapping(self, row: pd.Series, col_config: ColumnConfig) -> Any:
        """Aplicar mapeo dinámico a una fila específica"""
        try:
            # Obtener valor de la columna fuente
            if col_config.mapping_source not in row.index:
                return ''
            
            source_value = str(row[col_config.mapping_source]).strip()
            
            # Crear identificador único para este mapeo
            mapping_id = f"{col_config.mapping_source}_{col_config.mapping_key_column}_{col_config.mapping_value_column}"
            
            # Usar cache para mapeos
            cache_key = f"{mapping_id}_{source_value}"
            if cache_key in self._mapping_cache:
                return self._mapping_cache[cache_key]
            
            # Buscar en la configuración de mapeo
            if self.mapping_config and mapping_id in self.mapping_config:
                mapping_info = self.mapping_config[mapping_id]
                mapping_data = mapping_info['mapping_dict']
                additional_keys = mapping_info.get('additional_keys', [])
                
                # Crear clave de búsqueda
                if additional_keys:
                    # Clave compuesta con columnas adicionales
                    search_parts = [source_value]
                    
                    # Agregar valores de columnas adicionales del archivo fuente
                    for additional_col in additional_keys:
                        if additional_col in row.index:
                            search_parts.append(str(row[additional_col]).strip())
                        else:
                            search_parts.append("")
                    
                    search_key = "|".join(search_parts)
                else:
                    # Búsqueda simple
                    search_key = source_value
                
                # Buscar el valor en el mapeo con múltiples estrategias
                result = self._find_mapped_value(mapping_data, search_key, source_value)
                self._mapping_cache[cache_key] = result
                return result
            else:
                # Si no hay configuración de mapeo, devolver valor original
                return source_value
                
        except Exception as e:
            self.logger.error(f"Error aplicando mapeo para columna '{col_config.display_name}': {e}")
            return row.get(col_config.mapping_source, '')
    
    def _find_mapped_value(self, mapping_data: Dict, search_key: str, default_value: Any) -> Any:
        """Buscar valor en el diccionario de mapeo con diferentes estrategias"""
        # 1. Búsqueda exacta
        if search_key in mapping_data:
            return mapping_data[search_key]
        
        # 2. Búsqueda case-insensitive
        for key, value in mapping_data.items():
            if str(key).strip().lower() == search_key.lower():
                return value
        
        # 3. Búsqueda parcial (para casos donde hay espacios extra)
        search_key_clean = search_key.strip()
        for key, value in mapping_data.items():
            if str(key).strip() == search_key_clean:
                return value
        
        # 4. Búsqueda por similitud (para casos donde hay pequeñas diferencias)
        for key, value in mapping_data.items():
            key_str = str(key).strip()
            if len(key_str) > 3 and len(search_key) > 3:
                # Si más del 80% de los caracteres coinciden
                similarity = self._calculate_similarity(key_str, search_key)
                if similarity > 0.8:
                    return value
        
        # 5. Si no se encuentra, devolver valor por defecto
        return default_value
    
    def _calculate_similarity(self, str1: str, str2: str) -> float:
        """Calcular similitud entre dos strings"""
        try:
            from difflib import SequenceMatcher
            return SequenceMatcher(None, str1.lower(), str2.lower()).ratio()
        except:
            return 0.0
    
    def _normalize_date_value(self, date_value: str) -> str:
        """Normalizar valor de fecha a formato dd/mm/yy"""
        try:
            if not date_value or date_value.strip() == "":
                return ""
            
            # Si ya está en formato dd/mm/yy, devolver tal como está
            if len(date_value.split('/')) == 3 and len(date_value.split('/')[2]) == 2:
                return date_value
            
            # Intentar parsear diferentes formatos
            for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(date_value, fmt)
                    return dt.strftime('%d/%m/%y')
                except ValueError:
                    continue
            
            # Si no se puede parsear, devolver el valor original
            return date_value
            
        except Exception as e:
            self.logger.warning(f"Error normalizando fecha {date_value}: {e}")
            return date_value
    
    def create_preview_data(self, 
                      data: pd.DataFrame,
                      columns_config: List[ColumnConfig],
                      max_rows: int = None,
                      mapping_config: Dict = None) -> pd.DataFrame:
        """Crear datos de vista previa aplicando configuraciones de columnas"""
        try:
            # Configurar mapeo si se proporciona
            if mapping_config:
                self.mapping_config = mapping_config
            
            # Reiniciar contadores del generador numérico para preview
            if self.numeric_generator:
                self.numeric_generator.reset_counters()
            
            # Limpiar cache
            self._mapping_cache.clear()
            
            # Pre-procesar grupos numéricos para todas las columnas que lo necesiten
            for col_config in columns_config:
                if hasattr(col_config, 'is_numeric_generator') and col_config.is_numeric_generator:
                    self._preprocess_numeric_groups(data, col_config)
            
            # Limitar filas si se especifica
            preview_data = data.head(max_rows) if max_rows else data
            
            # Crear DataFrame de vista previa
            preview_df = pd.DataFrame()
            
            # Procesar cada columna configurada
            for col_config in columns_config:
                column_values = []
                
                for idx, row in preview_data.iterrows():
                    value = self._get_column_value(row, col_config, data)
                    formatted_value = self._format_value(value, col_config)
                    column_values.append(formatted_value)
                
                preview_df[col_config.display_name] = column_values
            
            return preview_df
            
        except Exception as e:
            self.logger.error(f"Error creando vista previa: {e}")
            return pd.DataFrame()
    
    def get_export_info(self, data: pd.DataFrame, column_configs: List[ColumnConfig]) -> Dict[str, Any]:
        """Obtener información sobre la exportación, incluyendo si se dividirá en múltiples archivos"""
        total_rows = len(data)
        max_rows_per_file = 10000
        will_split = total_rows > max_rows_per_file
        
        if will_split:
            num_files = (total_rows + max_rows_per_file - 1) // max_rows_per_file
            estimated_size_mb = (total_rows * len(column_configs) * 10) / (1024 * 1024)
            estimated_size_per_file_mb = estimated_size_mb / num_files
        else:
            num_files = 1
            estimated_size_mb = (total_rows * len(column_configs) * 10) / (1024 * 1024)
            estimated_size_per_file_mb = estimated_size_mb
        
        return {
            'total_rows': total_rows,
            'total_columns': len(column_configs),
            'will_split': will_split,
            'num_files': num_files,
            'max_rows_per_file': max_rows_per_file,
            'estimated_size_mb': round(estimated_size_mb, 2),
            'estimated_size_per_file_mb': round(estimated_size_per_file_mb, 2),
            'split_warning': f"El archivo será dividido en {num_files} partes debido a su tamaño ({total_rows:,} filas)" if will_split else None
        }
    
    def export_with_multiple_sheets(self,
                                  sheets_data: Dict[str, Dict[str, Any]],
                                  filename: str = None) -> bool:
        """Exportar archivo Excel con múltiples hojas"""
        try:
            # Generar nombre de archivo si no se proporciona
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"excel_export_multi_{timestamp}.xlsx"
            
            # Crear workbook
            self.workbook = Workbook()
            
            # Eliminar hoja por defecto
            self.workbook.remove(self.workbook.active)
            
            # Crear cada hoja
            for sheet_name, sheet_info in sheets_data.items():
                data = sheet_info['data']
                columns_config = sheet_info['columns_config']
                
                # Crear nueva hoja
                self.worksheet = self.workbook.create_sheet(title=sheet_name)
                
                # Aplicar configuración
                self._apply_column_configuration(data, columns_config)
            
            # Guardar archivo
            export_path = Path(self.settings.default_export_dir) / filename
            self.workbook.save(export_path)
            
            self.logger.info(f"Archivo Excel multi-hoja creado: {export_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creando archivo multi-hoja: {e}")
            return False
    
    def validate_export_data(self, 
                           data: pd.DataFrame,
                           columns_config: List[ColumnConfig]) -> List[str]:
        """Validar datos antes de exportar"""
        errors = []
        
        if data.empty:
            errors.append("No hay datos para exportar")
            return errors
        
        if not columns_config:
            errors.append("No hay configuración de columnas")
            return errors
        
        # Validar columnas requeridas
        for col_config in columns_config:
            if col_config.required and col_config.source_column:
                if col_config.source_column not in data.columns:
                    errors.append(f"Columna requerida no encontrada: {col_config.source_column}")
                elif data[col_config.source_column].isna().all():
                    errors.append(f"Columna requerida está vacía: {col_config.source_column}")
        
        return errors
    
    def get_export_statistics(self, 
                            data: pd.DataFrame,
                            columns_config: List[ColumnConfig]) -> Dict[str, Any]:
        """Obtener estadísticas de exportación"""
        return {
            'total_rows': len(data),
            'total_columns': len(columns_config),
            'source_columns': len([col for col in columns_config if col.source_column]),
            'generated_columns': len([col for col in columns_config if col.is_generated]),
            'required_columns': len([col for col in columns_config if col.required]),
            'data_types': {dt.value: len([col for col in columns_config if col.data_type == dt]) 
                          for dt in DataType},
            'estimated_file_size_mb': (len(data) * len(columns_config) * 10) / (1024 * 1024)  # Estimación aproximada
        }
    
    def cleanup(self):
        """Limpiar recursos"""
        self.workbook = None
        self.worksheet = None
        self._mapping_cache.clear()
        self._date_format_cache.clear()
        self._column_letter_cache.clear()
        self.logger.info("Recursos de exportación limpiados")
    
    def export_excel(self, source_data: pd.DataFrame, column_configs: List[ColumnConfig], 
                     numeric_config: Dict = None, mapping_config: Dict = None, 
                     export_config: Dict = None, progress_callback: Callable = None) -> Dict[str, Any]:
        """Exportar archivo Excel con configuración completa - Versión Optimizada"""
        try:
            # Almacenar configuración de mapeo para uso en _get_column_value
            self.mapping_config = mapping_config
            
            # Limpiar cache al inicio
            self._mapping_cache.clear()
            self._date_format_cache.clear()
            self._column_letter_cache.clear()
            
            if progress_callback:
                progress_callback(10)
            
            # Validar datos
            errors = self.validate_export_data(source_data, column_configs)
            if errors:
                raise ValueError(f"Errores de validación: {'; '.join(errors)}")
            
            if progress_callback:
                progress_callback(20)
            
            # Verificar si necesitamos dividir la exportación (más de 10,000 filas)
            max_rows_per_file = 10000
            total_rows = len(source_data)
            
            if total_rows > max_rows_per_file:
                self.logger.info(f"Archivo grande detectado ({total_rows} filas). Dividiendo en múltiples archivos...")
                return self._export_large_file(source_data, column_configs, export_config, max_rows_per_file, progress_callback)
            
            if progress_callback:
                progress_callback(30)
            
            # Configurar nombre de archivo
            output_file = export_config.get('output_file') if export_config else None
            if not output_file:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"excel_export_{timestamp}.xlsx"
            
            # Crear archivo Excel
            sheet_name = export_config.get('sheet_name', 'Datos') if export_config else 'Datos'
            success = self.create_excel_file(source_data, column_configs, output_file, sheet_name)
            
            if progress_callback:
                progress_callback(90)
            
            if not success:
                raise Exception("Error al crear el archivo Excel")
            
            if progress_callback:
                progress_callback(100)
            
            return {
                'success': True,
                'file_path': output_file,
                'rows_processed': len(source_data),
                'files_created': 1
            }
            
        except PermissionError as pe:
            error_msg = f"Error de permisos al exportar: {pe}"
            self.logger.error(error_msg)
            raise Exception(error_msg)
        except Exception as e:
            error_msg = f"Error en export_excel: {e}"
            self.logger.error(error_msg)
            raise Exception(error_msg)
    
    def _format_value(self, value: Any, col_config: ColumnConfig) -> Any:
        """Formatear valor según configuración de columna - Versión Optimizada"""
        if value is None or pd.isna(value):
            return ""
        
        # Aplicar formato según tipo de datos
        if col_config.data_type == DataType.NUMBER:
            try:
                num_value = float(value)
                
                # Aplicar formato específico si está configurado
                if hasattr(col_config, 'format_string') and col_config.format_string:
                    format_str = col_config.format_string.strip()
                    
                    # Cache para formatos de números
                    cache_key = f"num_{format_str}_{num_value}"
                    if cache_key in self._date_format_cache:
                        return self._date_format_cache[cache_key]
                    
                    # Manejar formatos específicos
                    if format_str == '#':
                        result = int(num_value) if num_value == int(num_value) else int(num_value)
                    elif format_str == '##':
                        result = int(num_value) if num_value == int(num_value) else int(num_value)
                    elif format_str == '0.00':
                        result = f"{num_value:.2f}"
                    elif format_str == '0.00%':
                        result = f"{num_value:.2%}"
                    elif format_str == '#,##0.00':
                        result = f"{num_value:,.2f}"
                    elif format_str == '#,##0':
                        result = f"{int(num_value):,}" if num_value == int(num_value) else f"{int(num_value):,}"
                    else:
                        # Intentar usar el formato como especificador de Python
                        try:
                            result = format(num_value, format_str)
                        except ValueError:
                            # Si falla, devolver como entero si es posible
                            result = int(num_value) if num_value == int(num_value) else num_value
                    
                    self._date_format_cache[cache_key] = result
                    return result
                else:
                    # Sin formato específico, devolver como entero si es posible
                    return int(num_value) if num_value == int(num_value) else num_value
                    
            except (ValueError, TypeError):
                return value
                
        elif col_config.data_type == DataType.DATE:
            if isinstance(value, (str, pd.Timestamp, datetime)):
                try:
                    # Cache para fechas
                    cache_key = f"date_{value}_{getattr(col_config, 'format_string', 'default')}"
                    if cache_key in self._date_format_cache:
                        return self._date_format_cache[cache_key]
                    
                    # Convertir a datetime si es necesario
                    if isinstance(value, str):
                        date_obj = None
                        date_formats = [
                            "%Y-%m-%d %H:%M:%S",  # 2025-02-18 00:00:00
                            "%Y-%m-%d",           # 2025-02-18
                            "%d/%m/%Y",           # 18/02/2025
                            "%d/%m/%y",           # 18/02/25
                            "%m/%d/%Y",           # 02/18/2025
                            "%m/%d/%y"            # 02/18/25
                        ]
                        
                        for fmt in date_formats:
                            try:
                                date_obj = datetime.strptime(value, fmt)
                                break
                            except ValueError:
                                continue
                                
                        if date_obj is None:
                            self._date_format_cache[cache_key] = value
                            return value
                    elif isinstance(value, pd.Timestamp):
                        date_obj = value.to_pydatetime()
                    else:
                        date_obj = value
                    
                    # Aplicar formato personalizado si está configurado
                    if hasattr(col_config, 'format_string') and col_config.format_string:
                        format_str = col_config.format_string.strip()
                        
                        # Mapear formatos comunes a strftime
                        format_mapping = {
                            'dd/mm/yy': '%d/%m/%y',
                            'dd/mm/yyyy': '%d/%m/%Y',
                            'yyyy-mm-dd': '%Y-%m-%d',
                            'yyyy-mm-dd hh:mm:ss': '%Y-%m-%d %H:%M:%S',
                            'mm/dd/yy': '%m/%d/%y',
                            'mm/dd/yyyy': '%m/%d/%Y'
                        }
                        
                        strftime_format = format_mapping.get(format_str.lower(), format_str)
                        
                        try:
                            result = date_obj.strftime(strftime_format)
                        except Exception as e:
                            result = date_obj.strftime('%d/%m/%y')  # Formato por defecto
                    else:
                        # Formato por defecto
                        result = date_obj.strftime('%d/%m/%y')
                    
                    self._date_format_cache[cache_key] = result
                    return result
                        
                except Exception as e:
                    return value
        
        return str(value)
    
    def create_excel_file(self, source_data: pd.DataFrame, column_configs: List[ColumnConfig], 
                         output_file: str, sheet_name: str = "Datos") -> bool:
        """Crear archivo Excel con datos y configuración - Versión Optimizada"""
        try:
            # Crear workbook
            self.workbook = Workbook()
            
            # Eliminar hoja por defecto
            self.workbook.remove(self.workbook.active)
            
            # Crear nueva hoja
            self.worksheet = self.workbook.create_sheet(title=sheet_name)
            
            # Aplicar configuración de columnas
            self._apply_column_configuration(source_data, column_configs)
            
            # Guardar archivo con manejo de conflictos
            export_path = Path(self.settings.default_export_dir) / output_file
            export_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Si el archivo existe, crear un nombre único
            if export_path.exists():
                base_name = export_path.stem
                extension = export_path.suffix
                counter = 1
                while export_path.exists():
                    new_name = f"{base_name}_{counter}{extension}"
                    export_path = Path(self.settings.default_export_dir) / new_name
                    counter += 1
                    if counter > 100:  # Evitar bucle infinito
                        raise Exception(f"No se pudo crear un nombre único para el archivo después de 100 intentos")
            
            # Intentar guardar el archivo
            try:
                self.workbook.save(export_path)
                self.logger.info(f"Archivo Excel creado: {export_path}")
                return True
            except PermissionError as pe:
                # Si hay error de permisos, intentar con un nombre temporal
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                temp_name = f"excel_export_{timestamp}.xlsx"
                temp_path = Path(self.settings.default_export_dir) / temp_name
                
                self.workbook.save(temp_path)
                self.logger.warning(f"Archivo guardado con nombre temporal debido a permisos: {temp_path}")
                return True
                
        except Exception as e:
            self.logger.error(f"Error creando archivo Excel: {e}")
            return False
    
    def _apply_column_configuration(self, data: pd.DataFrame, columns_config: List[ColumnConfig]):
        """Aplicar configuración de columnas al worksheet - Versión Optimizada"""
        try:
            # Reiniciar contadores del generador numérico
            if self.numeric_generator:
                self.numeric_generator.reset_counters()
            
            # Pre-procesar grupos numéricos para todas las columnas que lo necesiten
            for col_config in columns_config:
                if hasattr(col_config, 'is_numeric_generator') and col_config.is_numeric_generator:
                    self._preprocess_numeric_groups(data, col_config)
            
            # Crear encabezados
            headers = [col_config.display_name for col_config in columns_config]
            self.worksheet.append(headers)
            
            # Aplicar estilo a encabezados
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for col_idx, header in enumerate(headers, 1):
                cell = self.worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Procesar datos en lotes para optimizar memoria
            batch_size = 1000
            total_rows = len(data)
            
            for batch_start in range(0, total_rows, batch_size):
                batch_end = min(batch_start + batch_size, total_rows)
                batch_data = data.iloc[batch_start:batch_end]
                
                for row_idx, (_, row) in enumerate(batch_data.iterrows(), batch_start + 2):
                    for col_idx, col_config in enumerate(columns_config, 1):
                        value = self._get_column_value(row, col_config, data)
                        formatted_value = self._format_value(value, col_config)
                        
                        cell = self.worksheet.cell(row=row_idx, column=col_idx)
                        cell.value = formatted_value
                        
                        # Aplicar alineación según tipo de datos
                        if col_config.data_type == DataType.NUMBER:
                            cell.alignment = Alignment(horizontal="right")
                        elif col_config.data_type == DataType.DATE:
                            cell.alignment = Alignment(horizontal="center")
            
            # Ajustar ancho de columnas
            for col_idx, col_config in enumerate(columns_config, 1):
                # Cache para letras de columnas
                if col_idx not in self._column_letter_cache:
                    self._column_letter_cache[col_idx] = self.worksheet.cell(row=1, column=col_idx).column_letter
                
                column_letter = self._column_letter_cache[col_idx]
                self.worksheet.column_dimensions[column_letter].width = 15
                
        except Exception as e:
            self.logger.error(f"Error aplicando configuración de columnas: {e}")
            raise
    
    def _export_large_file(self, source_data: pd.DataFrame, column_configs: List[ColumnConfig], 
                          export_config: Dict, max_rows_per_file: int, progress_callback: Callable = None) -> Dict[str, Any]:
        """Exportar archivo grande dividido en múltiples partes - Versión Optimizada"""
        try:
            total_rows = len(source_data)
            num_files = (total_rows + max_rows_per_file - 1) // max_rows_per_file  # División hacia arriba
            
            self.logger.info(f"Dividiendo {total_rows} filas en {num_files} archivos de máximo {max_rows_per_file} filas cada uno")
            
            # Configurar nombre base del archivo
            base_output_file = export_config.get('output_file') if export_config else None
            if not base_output_file:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_output_file = f"excel_export_{timestamp}"
            else:
                # Remover extensión si existe
                base_output_file = Path(base_output_file).stem
            
            sheet_name = export_config.get('sheet_name', 'Datos') if export_config else 'Datos'
            created_files = []
            
            for file_index in range(num_files):
                if progress_callback:
                    # Calcular progreso basado en archivos procesados
                    progress = 20 + (file_index * 70 // num_files)
                    progress_callback(progress)
                
                # Calcular rango de filas para este archivo
                start_row = file_index * max_rows_per_file
                end_row = min((file_index + 1) * max_rows_per_file, total_rows)
                
                # Extraer datos para este archivo
                file_data = source_data.iloc[start_row:end_row].copy()
                
                # Crear nombre de archivo para esta parte
                if num_files == 1:
                    output_file = f"{base_output_file}.xlsx"
                else:
                    output_file = f"{base_output_file}_parte_{file_index + 1:03d}_de_{num_files:03d}.xlsx"
                
                self.logger.info(f"Creando archivo {file_index + 1}/{num_files}: {output_file} con {len(file_data)} filas")
                
                # Crear archivo Excel para esta parte
                success = self.create_excel_file(file_data, column_configs, output_file, sheet_name)
                
                if not success:
                    raise Exception(f"Error al crear el archivo {output_file}")
                
                created_files.append(output_file)
                
                # Limpiar recursos después de cada archivo para liberar memoria
                self.cleanup()
            
            if progress_callback:
                progress_callback(100)
            
            self.logger.info(f"Exportación completada. Se crearon {len(created_files)} archivos")
            
            return {
                'success': True,
                'file_path': created_files[0] if len(created_files) == 1 else f"{base_output_file}_parte_001_de_{num_files:03d}.xlsx",
                'rows_processed': total_rows,
                'files_created': len(created_files),
                'all_files': created_files,
                'split_info': {
                    'total_files': num_files,
                    'max_rows_per_file': max_rows_per_file,
                    'base_filename': base_output_file
                }
            }
            
        except Exception as e:
            error_msg = f"Error en exportación de archivo grande: {e}"
            self.logger.error(error_msg)
            raise Exception(error_msg)