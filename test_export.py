#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Prueba de exportación Excel con generador numérico
"""

import sys
import pandas as pd
from pathlib import Path
from datetime import datetime

# Agregar el directorio raíz al path para imports
root_dir = Path(__file__).parent
sys.path.insert(0, str(root_dir))

from core.export_manager import ExportManager
from config.settings import AppSettings
from models.column_config import ColumnConfig, DataType

def create_test_data():
    """Crear datos de prueba similares al ejemplo del usuario"""
    data = {
        'Nombres': [
            'YULIED AREVALO',
            'SANDRA MILENA VALENCIA MA',
            'LUIS EDUARDO AMAYA',
            'SILVIA ADRIANA RUANO CORTE',
            'LETICIA CERDAS AMADO',
            'SANDRA ANYELY ARIAS CÁCER',
            'MARIANA NUÑEZ CARDENAS',
            'KAROL MORENO CAICEDO',
            'MELINA ANDREA LASSO ORTEC'
        ],
        'Documento': ['60350189', '40331200', '1090400427', '27105646', '60334558', '37272258', '1108599321', '1022418751', '1085661803'],
        'Correo': ['asistentesgi@jucamal.com', 'sami1129@hotmail.es', 'asistentesgi@jucamal.com', 'sarc1384@gmail.com', 'saludocupacional@transmateria', 'anyely_07_@hotmail.com', 'imporberaca05@gmail.com', 'karol.moreno@proyecformas.cc', 'elin1332@gmail.com'],
        'Telefono': ['0', '0', '0', '0', '0', '0', '0', '0', '0'],
        'Nit': ['900409401.0', '892000435.0', '900409401.0', '901011594.0', '890500988.0', '900362666.0', '901489102.0', '830074655.0', '891201588.0'],
        'Razón Social': ['JUCAMAL SAS', 'COOPERATIVA DE TRANSPORT.', 'JUCAMAL SAS', 'GRUPO HERITAGE SAS', 'TRANSMATERIALES SA', 'CUCUTA MOTORS', 'IMPORTADORA FERRETERA BEF', 'PROYECFORMAS SAS', 'COACREMAT'],
        'Correo Empresa': ['asistentesgi@jucamal.com', 'Sgsst.cootransmeta01@gmail.co', 'asistentesgi@jucamal.com', 'sarc1384@gmail.com', 'saludocupacional@transmateria', 'anyely.arias@cucutamotors.con', 'imporberaca05@gmail.com', 'karol.moreno@proyecformas.cc', 'control.personal@coacremat.co'],
        'Teléfono Em': ['0', '0', '0', '0', '0', '0', '0', '0', '0'],
        'Ciudad': ['VIRTUAL', 'VIRTUAL', 'VIRTUAL', 'VIRTUAL', 'VIRTUAL', 'VIRTUAL', 'VIRTUAL', 'VIRTUAL', 'VIRTUAL'],
        'Fecha': ['2025-02-18 00:00:00', '2025-02-18 00:00:00', '2025-02-18 00:00:00', '2025-02-18 00:00:00', '2025-02-18 00:00:00', '2025-02-21 00:00:00', '2025-02-21 00:00:00', '2025-02-21 00:00:00', '2025-02-21 00:00:00'],
        'Módulo': ['Marco Legal', 'Marco Legal', 'Marco Legal', 'Marco Legal', 'Marco Legal', 'Sostenibilidad Empresarial', 'Sostenibilidad Empresarial', 'Sostenibilidad Empresarial', 'Sostenibilidad Empresarial'],
        'Tema': ['DECRETO 1072 DEL 2015: BASE FUNDAMENTAL DEL SG-SST', 'DECRETO 1072 DEL 2015: BASE FUNDAMENTAL DEL SG-SST', 'DECRETO 1072 DEL 2015: BASE FUNDAMENTAL DEL SG-SST', 'DECRETO 1072 DEL 2015: BASE FUNDAMENTAL DEL SG-SST', 'DECRETO 1072 DEL 2015: BASE FUNDAMENTAL DEL SG-SST', 'FUNCIONES Y RESPONSABILIDADES', 'FUNCIONES Y RESPONSABILIDADES', 'FUNCIONES Y RESPONSABILIDADES', 'FUNCIONES Y RESPONSABILIDADES'],
        'Estado': ['Asistio', 'Asistio', 'Asistio', 'Asistio', 'Asistio', 'Asistio', 'Asistio', 'Asistio', 'Asistio']
    }
    
    return pd.DataFrame(data)

def test_export():
    """Probar la exportación Excel con generador numérico"""
    print("🧪 PRUEBA DE EXPORTACIÓN EXCEL")
    print("=" * 60)
    
    # Crear datos de prueba
    data = create_test_data()
    print(f"Datos de prueba creados: {len(data)} filas")
    print()
    
    # Configurar export manager
    settings = AppSettings()
    export_manager = ExportManager(settings)
    
    # Crear configuración de columnas
    columns_config = [
        # Columna con generador numérico
        ColumnConfig(
            name='matricula',
            display_name='Matrícula',
            source_column='',
            data_type=DataType.TEXT,
            required=False,
            is_generated=True,
            is_numeric_generator=True,
            numeric_start=1,
            numeric_grouping_columns=['Fecha', 'Tema', 'Módulo']
        ),
        # Otras columnas
        ColumnConfig(
            name='nombres',
            display_name='Nombres',
            source_column='Nombres',
            data_type=DataType.TEXT,
            required=False,
            is_generated=False
        ),
        ColumnConfig(
            name='documento',
            display_name='Documento',
            source_column='Documento',
            data_type=DataType.TEXT,
            required=False,
            is_generated=False
        ),
        ColumnConfig(
            name='fecha',
            display_name='Fecha',
            source_column='Fecha',
            data_type=DataType.DATE,
            required=False,
            is_generated=False,
            format_string='dd/mm/yy'
        ),
        ColumnConfig(
            name='modulo',
            display_name='Módulo',
            source_column='Módulo',
            data_type=DataType.TEXT,
            required=False,
            is_generated=False
        ),
        ColumnConfig(
            name='tema',
            display_name='Tema',
            source_column='Tema',
            data_type=DataType.TEXT,
            required=False,
            is_generated=False
        )
    ]
    
    # Configuración de exportación
    export_config = {
        'output_file': 'test_export_numeric.xlsx',
        'sheet_name': 'Datos'
    }
    
    print("📝 EXPORTANDO ARCHIVO EXCEL...")
    print("-" * 40)
    
    try:
        # Exportar archivo
        result = export_manager.export_excel(
            source_data=data,
            column_configs=columns_config,
            export_config=export_config
        )
        
        if result['success']:
            print(f"✅ Archivo exportado exitosamente: {result['file_path']}")
            print(f"📊 Filas procesadas: {result['rows_processed']}")
            
            # Verificar el archivo exportado
            import openpyxl
            # Buscar el archivo en la carpeta exportados
            export_path = Path("exportados") / "test_export_numeric.xlsx"
            wb = openpyxl.load_workbook(export_path)
            ws = wb.active
            
            print("\n📋 CONTENIDO DEL ARCHIVO EXPORTADO:")
            print("-" * 40)
            
            # Leer las primeras filas para verificar
            for row_idx in range(1, min(11, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                
                if row_idx == 1:
                    print(f"Encabezados: {' | '.join(row_data)}")
                else:
                    print(f"Fila {row_idx}: {' | '.join(row_data)}")
            
            # Verificar números únicos en la columna Matrícula
            matricula_values = []
            for row_idx in range(2, ws.max_row + 1):
                matricula_value = ws.cell(row=row_idx, column=1).value
                if matricula_value is not None:
                    matricula_values.append(str(matricula_value))
            
            unique_matriculas = set(matricula_values)
            print(f"\n🔢 NÚMEROS ÚNICOS EN MATRÍCULA: {sorted(unique_matriculas)}")
            print(f"📊 Total de números únicos: {len(unique_matriculas)}")
            
            if len(unique_matriculas) == 2:
                print("✅ ÉXITO: Se generaron 2 números diferentes para 2 grupos")
            else:
                print(f"❌ ERROR: Se esperaban 2 grupos, pero se generaron {len(unique_matriculas)} números únicos")
            
            wb.close()
            
        else:
            print("❌ Error en la exportación")
            
    except Exception as e:
        print(f"❌ Error durante la exportación: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_export()
